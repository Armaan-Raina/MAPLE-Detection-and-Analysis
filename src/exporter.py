"""
Excel exporter for validated MEA population events.

Produces a workbook with:
  Metadata    — recording info, detection parameters, model info
  Channel X   — one plain-data sheet per channel that has validated events
"""

import os
import warnings
from datetime import datetime

import numpy as np
from scipy import signal as sps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import mea_io
import detection


DECIMATION = 25


# ======================================================================
# Feature measurement
# ======================================================================

def _measure_features_for_channel(data, event_indices):
    """
    Return a list of feature dicts for each sample index.
    Filters once; computes morphological + spectral features + prominence per event.
    Returns None for an event if measurement fails.
    """
    filtered = detection.bandpass_filter(data)

    results = []
    for sample_idx in event_indices:
        try:
            feats = detection.compute_features(data, sample_idx)
        except (IndexError, ValueError):
            results.append(None)
            continue

        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                proms, _, _ = sps.peak_prominences(-filtered, [sample_idx])
            feats["prominence_uv"] = round(float(proms[0]), 3)
        except (ValueError, IndexError):
            feats["prominence_uv"] = None

        results.append(feats)

    return results


# ======================================================================
# Formatting constants (used only by the Metadata sheet)
# ======================================================================

_SECTION_FILL = PatternFill("solid", fgColor="1E4D8C")
_SECTION_FONT = Font(bold=True, color="FFFFFF")
_CENTER       = Alignment(horizontal="center", vertical="center")
_ALT_FILL     = PatternFill("solid", fgColor="EEF3FB")


def _fmt(val, decimals=3):
    if val is None:
        return "N/A"
    return round(float(val), decimals)


# ======================================================================
# Sheet writers
# ======================================================================

def _write_metadata(ws, session, recording_filename):
    n_analyzed    = len(session.channels)
    n_with_events = sum(1 for ch in session.channels.values() if ch.get("validated_events"))
    total_events  = sum(len(ch.get("validated_events", [])) for ch in session.channels.values())

    rows = [
        ("RECORDING",                    None),
        ("File",                         os.path.basename(recording_filename)),
        ("Export Date",                  datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("CHANNELS",                     None),
        ("Channels Analyzed",            n_analyzed),
        ("Channels with Events",         n_with_events),
        ("Total Validated Events",       total_events),
    ]

    for row_idx, (key, val) in enumerate(rows, start=1):
        if val is None:
            c           = ws.cell(row=row_idx, column=1, value=key)
            c.font      = _SECTION_FONT
            c.fill      = _SECTION_FILL
            c.alignment = _CENTER
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx,   end_column=2,
            )
        else:
            k_cell       = ws.cell(row=row_idx, column=1, value=key)
            v_cell       = ws.cell(row=row_idx, column=2, value=val)
            k_cell.font  = Font(bold=True)
            if row_idx % 2 == 0:
                k_cell.fill = _ALT_FILL
                v_cell.fill = _ALT_FILL

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32


def _write_channel_sheet(ws, events, features, comments, fs):
    """Write one channel's events as plain data — no fill, no bold, no color."""
    ws.append([
        "Event #", "Time (s)", "IEI (s)",
        "Amplitude (µV)", "Prominence (µV)",
        "Duration (ms)", "Rise Time (ms)", "Decay Time (ms)", "Half-Width (ms)",
        "Slope (µV/ms)", "Area (µV·ms)",
        "Comment",
    ])

    prev_time_s = None
    for ev_num, (sample_idx, feats) in enumerate(zip(events, features), start=1):
        time_s  = round(sample_idx / fs, 4)
        iei     = round(time_s - prev_time_s, 4) if prev_time_s is not None else "N/A"
        prev_time_s = time_s
        comment = comments.get(str(sample_idx), "")

        if feats:
            row = [
                ev_num, time_s, iei,
                _fmt(feats["amplitude_uv"]),
                _fmt(feats.get("prominence_uv")),
                _fmt(feats["duration_s"]          * 1000, 4),
                _fmt(feats["rise_time_s"]         * 1000, 4),
                _fmt(feats["decay_time_s"]        * 1000, 4),
                _fmt(feats["half_width_s"]        * 1000, 4),
                _fmt(feats["slope_uv_per_ms"]),
                _fmt(feats["area_uv_ms"], 1),
                comment,
            ]
        else:
            row = [ev_num, time_s, iei] + ["N/A"] * 8 + [comment]

        ws.append(row)


# ======================================================================
# Public entry point
# ======================================================================

def export_events_excel(path, session, recording_filename, fs):
    """Write the Excel workbook: Metadata sheet + one plain sheet per channel."""
    wb = openpyxl.Workbook()

    ws_meta = wb.active
    ws_meta.title = "Metadata"
    _write_metadata(ws_meta, session, recording_filename)

    for key in sorted(session.channels.keys(), key=int):
        ch     = session.channels[key]
        label  = ch["label"]
        ch_idx = int(key)

        events   = sorted(ch.get("validated_events", []))
        comments = ch.get("comments", {})
        if not events:
            continue

        try:
            data, _ = mea_io.read_and_decimate(
                recording_filename, ch_idx, decimation=DECIMATION
            )
            features = _measure_features_for_channel(data, events)
        except (IndexError, OSError):
            features = [None] * len(events)

        ws = wb.create_sheet(f"Channel {label}")
        _write_channel_sheet(ws, events, features, comments, fs)

    wb.save(path)
